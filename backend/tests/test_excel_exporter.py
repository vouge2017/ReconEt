"""
Tests for Excel Export Engine

Verifies:
- Excel file generation
- Correct sheets created
- Data integrity
- Styling applied
"""

import os
import io
import pytest

import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

from app.engine.excel_exporter import ExcelExporter, ExportConfig, export_reconciliation


@pytest.fixture
def exporter():
    return ExcelExporter()


@pytest.fixture
def sample_transactions():
    """Create sample transactions for testing"""
    from dataclasses import dataclass
    
    @dataclass
    class Txn:
        row_index: int = 0
        date: str = ""
        value_date: str = ""
        particulars: str = ""
        reference: str = ""
        narrative: str = ""
        debit: float = 0.0
        credit: float = 0.0
        balance: float = 0.0
        fee_amount: float = 0.0
        bank_charge: float = 0.0
        gov_tax: float = 0.0
        reference_code: str = ""
        transaction_type: str = ""
        cheque_number: str = ""
        is_matched: bool = False
        gross_amount: float = 0.0
        net_amount: float = 0.0
    
    return [
        Txn(row_index=1, date="01/01/2022", particulars="ATM Cash Withdrawal",
            reference="FT22001773M3", narrative="ATM Cash Withdrawal",
            debit=1002.0, balance=4125.54, fee_amount=2.0,
            bank_charge=1.74, gov_tax=0.26, reference_code="FT"),
        Txn(row_index=2, date="03/01/2022", particulars="ATM Cash Withdrawal",
            reference="FT2200357RZF", narrative="ATM Cash Withdrawal",
            debit=1002.0, balance=3123.54, fee_amount=2.0,
            bank_charge=1.74, gov_tax=0.26, reference_code="FT"),
        Txn(row_index=3, date="06/01/2022", particulars="Transfer",
            reference="FT220063HFM4", narrative="Transfer",
            credit=10000.0, balance=11221.14, reference_code="FT"),
    ]


@pytest.fixture
def sample_config():
    return ExportConfig(
        company_name="Test Company",
        account_number="1000066001079",
        period="01 JAN 2022 - 05 APR 2022",
    )


class TestExcelExporter:
    """Test Excel export functionality"""
    
    def test_export_returns_bytesio(self, exporter, sample_transactions, sample_config):
        """Export should return a BytesIO object"""
        result = exporter.export(sample_transactions, config=sample_config)
        assert isinstance(result, io.BytesIO)
    
    def test_export_is_valid_xlsx(self, exporter, sample_transactions, sample_config):
        """Export should be a valid .xlsx file"""
        try:
            import openpyxl
            buffer = exporter.export(sample_transactions, config=sample_config)
            wb = openpyxl.load_workbook(buffer)
            assert wb is not None
            wb.close()
        except ImportError:
            pytest.skip("openpyxl not installed")
    
    def test_export_has_summary_sheet(self, exporter, sample_transactions, sample_config):
        """Export should have Summary sheet"""
        import openpyxl
        buffer = exporter.export(sample_transactions, config=sample_config)
        wb = openpyxl.load_workbook(buffer)
        assert 'Summary' in wb.sheetnames
        wb.close()
    
    def test_export_has_transactions_sheet(self, exporter, sample_transactions, sample_config):
        """Export should have Bank Transactions sheet"""
        import openpyxl
        buffer = exporter.export(sample_transactions, config=sample_config)
        wb = openpyxl.load_workbook(buffer)
        assert 'Bank Transactions' in wb.sheetnames
        wb.close()
    
    def test_summary_contains_company_name(self, exporter, sample_transactions, sample_config):
        """Summary sheet should contain company name"""
        import openpyxl
        buffer = exporter.export(sample_transactions, config=sample_config)
        wb = openpyxl.load_workbook(buffer)
        ws = wb['Summary']
        
        found = False
        for row in ws.iter_rows(values_only=True):
            if row[0] == 'Company' and row[1] == 'Test Company':
                found = True
                break
        assert found, "Company name not found in Summary"
        wb.close()
    
    def test_transaction_count(self, exporter, sample_transactions, sample_config):
        """Bank Transactions sheet should have correct number of rows"""
        import openpyxl
        buffer = exporter.export(sample_transactions, config=sample_config)
        wb = openpyxl.load_workbook(buffer)
        ws = wb['Bank Transactions']
        
        # Count data rows (excluding header)
        data_rows = ws.max_row - 1
        assert data_rows == len(sample_transactions)
        wb.close()
    
    def test_empty_transactions(self, exporter, sample_config):
        """Export should handle empty transaction list"""
        result = exporter.export([], config=sample_config)
        assert isinstance(result, io.BytesIO)
    
    def test_export_with_none_config(self, exporter, sample_transactions):
        """Export should work with default config"""
        result = exporter.export(sample_transactions, config=None)
        assert isinstance(result, io.BytesIO)


class TestExportConfig:
    """Test export configuration"""
    
    def test_default_config(self):
        """Default config should include all sections"""
        config = ExportConfig()
        assert config.include_matched is True
        assert config.include_unmatched_bank is True
        assert config.include_summary is True
        assert config.include_fee_breakdown is True
    
    def test_custom_config(self):
        """Custom config should be respected"""
        config = ExportConfig(
            company_name="My Company",
            include_matched=False,
        )
        assert config.company_name == "My Company"
        assert config.include_matched is False


class TestExportConvenienceFunction:
    """Test the convenience export function"""
    
    def test_export_reconciliation(self, sample_transactions):
        """Convenience function should work"""
        result = export_reconciliation(sample_transactions)
        assert isinstance(result, io.BytesIO)


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
