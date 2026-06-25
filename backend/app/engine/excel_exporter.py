"""
Excel Export Engine — Export reconciliation results to .xlsx

Generates professional Excel reports that accountants can:
- Share with their team
- Import into their ERP
- Present to CFOs
- Submit to auditors

Output includes:
- Matched transactions (with fee breakdown)
- Unmatched bank transactions
- Unmatched GL entries
- Summary statistics
- Exception categories
"""

import io
from datetime import datetime
from typing import List, Dict, Optional, Any
from dataclasses import dataclass


@dataclass
class ExportConfig:
    """Configuration for Excel export"""
    include_matched: bool = True
    include_unmatched_bank: bool = True
    include_unmatched_gl: bool = True
    include_summary: bool = True
    include_fee_breakdown: bool = True
    include_exceptions: bool = True
    company_name: str = ""
    account_number: str = ""
    period: str = ""


class ExcelExporter:
    """
    Export reconciliation results to Excel.
    
    Usage:
        exporter = ExcelExporter()
        buffer = exporter.export(transactions, config)
        # buffer is a BytesIO object ready for download
    """
    
    # Column definitions for each sheet
    MATCHED_COLUMNS = [
        ("Row", 6),
        ("Date", 12),
        ("Value Date", 12),
        ("Particulars", 20),
        ("Reference", 20),
        ("Narrative", 35),
        ("Bank Debit", 14),
        ("Bank Credit", 14),
        ("Bank Balance", 14),
        ("GL Entry", 25),
        ("GL Debit", 14),
        ("GL Credit", 14),
        ("Match Type", 15),
        ("Confidence", 12),
        ("Fee Amount", 12),
        ("Bank Charge", 12),
        ("Gov Tax", 12),
        ("WHT", 12),
        ("Gross Amount", 14),
        ("Net Amount", 14),
        ("Explanation", 40),
    ]
    
    UNMATCHED_BANK_COLUMNS = [
        ("Row", 6),
        ("Date", 12),
        ("Value Date", 12),
        ("Particulars", 20),
        ("Reference", 20),
        ("Narrative", 35),
        ("Debit", 14),
        ("Credit", 14),
        ("Balance", 14),
        ("Fee Amount", 12),
        ("Possible Reason", 30),
    ]
    
    UNMATCHED_GL_COLUMNS = [
        ("Entry ID", 15),
        ("Date", 12),
        ("Account Code", 15),
        ("Account Name", 25),
        ("Description", 35),
        ("Debit", 14),
        ("Credit", 14),
        ("Reference", 20),
        ("Possible Reason", 30),
    ]
    
    SUMMARY_COLUMNS = [
        ("Metric", 30),
        ("Value", 20),
    ]
    
    def export(
        self,
        transactions: List[Any],
        config: Optional[ExportConfig] = None,
        unmatched_bank: Optional[List[Any]] = None,
        unmatched_gl: Optional[List[Any]] = None,
        match_results: Optional[List[Any]] = None,
    ) -> io.BytesIO:
        """
        Export reconciliation results to Excel.
        
        Args:
            transactions: List of parsed bank transactions
            config: Export configuration
            unmatched_bank: Unmatched bank transactions (if available)
            unmatched_gl: Unmatched GL entries (if available)
            match_results: Matching results (if available)
        
        Returns:
            BytesIO object containing the Excel file
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        if config is None:
            config = ExportConfig()
        
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        currency_format = '#,##0.00'
        date_format = 'DD/MM/YYYY'
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Alternating row colors
        even_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        
        def style_header(ws, columns, row=1):
            """Apply header styling"""
            for col_idx, (name, width) in enumerate(columns, 1):
                cell = ws.cell(row=row, column=col_idx, value=name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        def style_cell(cell, row_idx, is_currency=False, is_date=False):
            """Apply cell styling"""
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = even_fill
            if is_currency and isinstance(cell.value, (int, float)):
                cell.number_format = currency_format
            if is_date and cell.value:
                cell.number_format = date_format
            cell.alignment = Alignment(vertical="center")
        
        # === Sheet 1: Summary ===
        if config.include_summary:
            ws_summary = wb.active
            ws_summary.title = "Summary"
            style_header(ws_summary, self.SUMMARY_COLUMNS)
            
            summary_data = self._build_summary(transactions, match_results, config)
            for row_idx, (metric, value) in enumerate(summary_data, 2):
                ws_summary.cell(row=row_idx, column=1, value=metric)
                ws_summary.cell(row=row_idx, column=2, value=value)
                style_cell(ws_summary.cell(row=row_idx, column=1), row_idx)
                style_cell(ws_summary.cell(row=row_idx, column=2), row_idx)
        
        # === Sheet 2: Matched Transactions ===
        if config.include_matched and match_results:
            ws_matched = wb.create_sheet("Matched Transactions")
            style_header(ws_matched, self.MATCHED_COLUMNS)
            
            for row_idx, match in enumerate(match_results, 2):
                self._write_match_row(ws_matched, row_idx, match, style_cell)
        
        # === Sheet 3: All Bank Transactions ===
        ws_all = wb.create_sheet("Bank Transactions")
        all_cols = [
            ("Row", 6), ("Date", 12), ("Value Date", 12),
            ("Particulars", 20), ("Reference", 20), ("Narrative", 35),
            ("Debit", 14), ("Credit", 14), ("Balance", 14),
            ("Fee Amount", 12), ("Bank Charge", 12), ("Gov Tax", 12),
            ("Reference Code", 12), ("Transaction Type", 20),
            ("Cheque Number", 15), ("Is Matched", 10),
        ]
        style_header(ws_all, all_cols)
        
        for row_idx, txn in enumerate(transactions, 2):
            self._write_transaction_row(ws_all, row_idx, txn, style_cell)
        
        # === Sheet 4: Unmatched Bank ===
        if config.include_unmatched_bank and unmatched_bank:
            ws_unmatched = wb.create_sheet("Unmatched Bank")
            style_header(ws_unmatched, self.UNMATCHED_BANK_COLUMNS)
            
            for row_idx, txn in enumerate(unmatched_bank, 2):
                self._write_unmatched_bank_row(ws_unmatched, row_idx, txn, style_cell)
        
        # === Sheet 5: Unmatched GL ===
        if config.include_unmatched_gl and unmatched_gl:
            ws_gl = wb.create_sheet("Unmatched GL")
            style_header(ws_gl, self.UNMATCHED_GL_COLUMNS)
            
            for row_idx, entry in enumerate(unmatched_gl, 2):
                self._write_unmatched_gl_row(ws_gl, row_idx, entry, style_cell)
        
        # === Sheet 6: Exceptions ===
        if config.include_exceptions:
            ws_exceptions = wb.create_sheet("Exceptions")
            exc_cols = [
                ("Category", 20), ("Count", 10), ("Total Amount", 14),
                ("Description", 40), ("Suggested Action", 40),
            ]
            style_header(ws_exceptions, exc_cols)
            
            exceptions = self._build_exceptions(transactions, unmatched_bank, match_results)
            for row_idx, exc in enumerate(exceptions, 2):
                for col_idx, val in enumerate(exc, 1):
                    cell = ws_exceptions.cell(row=row_idx, column=col_idx, value=val)
                    style_cell(cell, row_idx, is_currency=(col_idx == 3))
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def _build_summary(
        self, transactions, match_results, config
    ) -> List[tuple]:
        """Build summary statistics"""
        total_txns = len(transactions) if transactions else 0
        total_matched = len(match_results) if match_results else 0
        total_unmatched = total_txns - total_matched
        match_rate = (total_matched / total_txns * 100) if total_txns > 0 else 0
        
        total_debit = sum(getattr(t, 'debit', 0) or 0 for t in (transactions or []))
        total_credit = sum(getattr(t, 'credit', 0) or 0 for t in (transactions or []))
        total_fees = sum(getattr(t, 'fee_amount', 0) or 0 for t in (transactions or []))
        total_bank_charge = sum(getattr(t, 'bank_charge', 0) or 0 for t in (transactions or []))
        total_tax = sum(getattr(t, 'gov_tax', 0) or 0 for t in (transactions or []))
        
        summary = [
            ("Reconciliation Summary", ""),
            ("", ""),
            ("Company", config.company_name or "N/A"),
            ("Account", config.account_number or "N/A"),
            ("Period", config.period or "N/A"),
            ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("", ""),
            ("Transaction Statistics", ""),
            ("Total Bank Transactions", total_txns),
            ("Matched", total_matched),
            ("Unmatched", total_unmatched),
            ("Match Rate", f"{match_rate:.1f}%"),
            ("", ""),
            ("Amount Summary", ""),
            ("Total Debits", total_debit),
            ("Total Credits", total_credit),
            ("Net Movement", total_credit - total_debit),
            ("", ""),
            ("Fee Summary", ""),
            ("Total Fees", total_fees),
            ("Bank Charges", total_bank_charge),
            ("Government Tax (VAT)", total_tax),
        ]
        
        return summary
    
    def _write_match_row(self, ws, row_idx, match, style_cell):
        """Write a single matched transaction row"""
        txn = getattr(match, 'bank_transaction', match)
        gl = getattr(match, 'gl_entry', None)
        
        values = [
            getattr(txn, 'row_index', row_idx - 1),
            str(getattr(txn, 'date', '')),
            str(getattr(txn, 'value_date', '')),
            getattr(txn, 'particulars', ''),
            getattr(txn, 'reference', ''),
            getattr(txn, 'narrative', ''),
            getattr(txn, 'debit', 0) or None,
            getattr(txn, 'credit', 0) or None,
            getattr(txn, 'balance', 0) or None,
            getattr(gl, 'description', '') if gl else '',
            getattr(gl, 'debit', 0) or None if gl else None,
            getattr(gl, 'credit', 0) or None if gl else None,
            getattr(match, 'match_type', ''),
            getattr(match, 'confidence', 0),
            getattr(txn, 'fee_amount', 0) or None,
            getattr(txn, 'bank_charge', 0) or None,
            getattr(txn, 'gov_tax', 0) or None,
            0,  # WHT placeholder
            getattr(txn, 'gross_amount', 0) or None,
            getattr(txn, 'net_amount', 0) or None,
            getattr(match, 'explanation', ''),
        ]
        
        currency_cols = {7, 8, 9, 11, 12, 15, 16, 17, 18, 19, 20}
        date_cols = {2, 3}
        
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            style_cell(
                cell, row_idx,
                is_currency=(col_idx in currency_cols),
                is_date=(col_idx in date_cols)
            )
    
    def _write_transaction_row(self, ws, row_idx, txn, style_cell):
        """Write a single bank transaction row"""
        values = [
            getattr(txn, 'row_index', row_idx - 1),
            str(getattr(txn, 'date', '')),
            str(getattr(txn, 'value_date', '')),
            getattr(txn, 'particulars', ''),
            getattr(txn, 'reference', ''),
            getattr(txn, 'narrative', ''),
            getattr(txn, 'debit', 0) or None,
            getattr(txn, 'credit', 0) or None,
            getattr(txn, 'balance', 0) or None,
            getattr(txn, 'fee_amount', 0) or None,
            getattr(txn, 'bank_charge', 0) or None,
            getattr(txn, 'gov_tax', 0) or None,
            getattr(txn, 'reference_code', ''),
            getattr(txn, 'transaction_type', ''),
            getattr(txn, 'cheque_number', ''),
            "Yes" if getattr(txn, 'is_matched', False) else "No",
        ]
        
        currency_cols = {7, 8, 9, 10, 11, 12}
        date_cols = {2, 3}
        
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            style_cell(
                cell, row_idx,
                is_currency=(col_idx in currency_cols),
                is_date=(col_idx in date_cols)
            )
    
    def _write_unmatched_bank_row(self, ws, row_idx, txn, style_cell):
        """Write an unmatched bank transaction row"""
        values = [
            getattr(txn, 'row_index', row_idx - 1),
            str(getattr(txn, 'date', '')),
            str(getattr(txn, 'value_date', '')),
            getattr(txn, 'particulars', ''),
            getattr(txn, 'reference', ''),
            getattr(txn, 'narrative', ''),
            getattr(txn, 'debit', 0) or None,
            getattr(txn, 'credit', 0) or None,
            getattr(txn, 'balance', 0) or None,
            getattr(txn, 'fee_amount', 0) or None,
            self._guess_unmatched_reason(txn),
        ]
        
        currency_cols = {7, 8, 9, 10}
        date_cols = {2, 3}
        
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            style_cell(
                cell, row_idx,
                is_currency=(col_idx in currency_cols),
                is_date=(col_idx in date_cols)
            )
    
    def _write_unmatched_gl_row(self, ws, row_idx, entry, style_cell):
        """Write an unmatched GL entry row"""
        values = [
            getattr(entry, 'id', ''),
            str(getattr(entry, 'date', '')),
            getattr(entry, 'account_code', ''),
            getattr(entry, 'account_name', ''),
            getattr(entry, 'description', ''),
            getattr(entry, 'debit', 0) or None,
            getattr(entry, 'credit', 0) or None,
            getattr(entry, 'reference', ''),
            "No matching bank transaction",
        ]
        
        currency_cols = {6, 7}
        date_cols = {2}
        
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            style_cell(
                cell, row_idx,
                is_currency=(col_idx in currency_cols),
                is_date=(col_idx in date_cols)
            )
    
    def _guess_unmatched_reason(self, txn) -> str:
        """Guess why a bank transaction didn't match"""
        narrative = (getattr(txn, 'narrative', '') or '').upper()
        particulars = (getattr(txn, 'particulars', '') or '').upper()
        combined = f"{narrative} {particulars}"
        
        if any(kw in combined for kw in ['INTEREST', 'TAX AMOUNT']):
            return "Interest/Tax — may not have GL entry"
        if any(kw in combined for kw in ['ATM', 'CASH WITHDRAWAL']):
            return "Cash withdrawal — may be petty cash"
        if any(kw in combined for kw in ['FEE', 'CHARGE', 'COMMISSION']):
            return "Bank fee — may be recorded separately"
        if getattr(txn, 'fee_amount', 0) > 0:
            return "Fee-embedded amount — check gross vs net"
        if any(kw in combined for kw in ['TRANSFER', 'FT']):
            return "Transfer — check if intercompany"
        return "Review manually"
    
    def _build_exceptions(self, transactions, unmatched_bank, match_results) -> List[tuple]:
        """Build exception summary"""
        exceptions = []
        
        # Count by category
        fee_txns = [t for t in (transactions or []) if (getattr(t, 'fee_amount', 0) or 0) > 0]
        interest_txns = [t for t in (transactions or []) 
                        if 'INTEREST' in (getattr(t, 'narrative', '') or '').upper()]
        atm_txns = [t for t in (transactions or []) 
                   if 'ATM' in (getattr(t, 'narrative', '') or '').upper()
                   or 'CASH WITHDRAWAL' in (getattr(t, 'narrative', '') or '').upper()]
        
        if fee_txns:
            total_fees = sum(getattr(t, 'fee_amount', 0) for t in fee_txns)
            exceptions.append((
                "Fee Transactions", len(fee_txns), total_fees,
                "Transactions with embedded bank fees",
                "Verify fee extraction accuracy"
            ))
        
        if interest_txns:
            total_interest = sum(getattr(t, 'credit', 0) or 0 for t in interest_txns)
            exceptions.append((
                "Interest/Income", len(interest_txns), total_interest,
                "Interest credits and tax deductions",
                "Confirm with GL interest income account"
            ))
        
        if atm_txns:
            total_atm = sum(getattr(t, 'debit', 0) or 0 for t in atm_txns)
            exceptions.append((
                "ATM/Cash", len(atm_txns), total_atm,
                "ATM withdrawals and cash transactions",
                "Reconcile with petty cash account"
            ))
        
        if unmatched_bank:
            total_unmatched = sum(
                (getattr(t, 'debit', 0) or 0) + (getattr(t, 'credit', 0) or 0) 
                for t in unmatched_bank
            )
            exceptions.append((
                "Unmatched", len(unmatched_bank), total_unmatched,
                "Bank transactions with no GL match",
                "Review and create manual journal entries"
            ))
        
        return exceptions


def export_reconciliation(
    transactions: List[Any],
    match_results: Optional[List[Any]] = None,
    config: Optional[ExportConfig] = None,
) -> io.BytesIO:
    """
    Convenience function: Export reconciliation results to Excel.
    
    Args:
        transactions: Parsed bank transactions
        match_results: Matching results (optional)
        config: Export configuration (optional)
    
    Returns:
        BytesIO object containing .xlsx file
    """
    exporter = ExcelExporter()
    return exporter.export(
        transactions=transactions,
        config=config,
        match_results=match_results,
    )
